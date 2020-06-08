"""
    Licensed under the Apache License, Version 2.0 (the "License");
    you may not use this file except in compliance with the License.
    You may obtain a copy of the License at

        http://www.apache.org/licenses/LICENSE-2.0

    Unless required by applicable law or agreed to in writing, software
    distributed under the License is distributed on an "AS IS" BASIS,
    WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
    See the License for the specific language governing permissions and
    limitations under the License.
"""
import csv
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.writer.excel import save_virtual_workbook
from collections import OrderedDict

from django.http import HttpResponse, HttpResponseRedirect, StreamingHttpResponse
from django.core.exceptions import PermissionDenied
from rest_framework.decorators import api_view

from gwells.roles import REGISTRIES_VIEWER_ROLE

from .views import person_search_qs

REGISTRY_EXPORT_HEADER_COLUMNS = [
    'person_name',
    'registration_number',
    'registration_status',
    'well_driller_orcs_number',
    'pump_installer_orcs_number',
    'description',
    'primary_certificate',
    'company_name',
    'company_address',
    'company_province_state',
    'company_tel',
    'company_fax',
    'company_email',
    'company_url',
    'contact_cell',
    'contact_tel',
    'contact_email',
    'person_guid',
    'org_guid',
]

def build_record(person, registration=None, organization=None, application=None):
    # NOTE: must be kept in sync with REGISTRY_EXPORT_HEADER_COLUMNS above
    record = {
        'person_name': person.name,
        'registration_number': registration.registration_no if registration else None,
        'registration_status': application.display_status if application else None,
        'well_driller_orcs_number': person.well_driller_orcs_no,
        'pump_installer_orcs_number': person.pump_installer_orcs_no,
        'description': application.subactivity.description if application else None,
        'primary_certificate': application.primary_certificate if application else None,
        'company_name': organization.name if organization else None,
        'company_address': organization.mailing_address if organization else None,
        'company_province_state': organization.province_state if organization else None,
        'company_tel': organization.main_tel if organization else None,
        'company_fax': organization.fax_tel if organization else None,
        'company_email': organization.email if organization else None,
        'company_url': organization.website_url if organization else None,
        'contact_cell': person.contact_cell,
        'contact_tel': person.contact_tel,
        'contact_email': person.contact_email,
        'person_guid': person.person_guid,
        'org_guid': registration.organization_id if registration else None,
    }

    return list(record.values())

def build_row(queryset):
    for person in queryset:
        registrantions = person.registrations.all()
        if len(registrantions) == 0:
            yield build_record(person)
        else:
            for registration in registrantions:
                applications = registration.applications.all()
                if len(applications) == 0:
                    yield build_record(person, registration, registration.organization)
                else:
                    for application in applications:
                        yield build_record(person, registration, registration.organization, application)

def ordered_person_search_qs(request):
    qs = person_search_qs(request)

    order_by = request.GET.get('ordering', None)
    if order_by and order_by.strip('-') in ['surname']: # check ordering param in whitelist
        qs = qs.order_by(order_by)
    else:
        qs = qs.order_by('surname')

    return qs

@api_view(['GET'])
def csv_export_v2(request):
    """
    Export the registry as CSV. This is done in a vanilla functional Django view instead
    of DRF, because DRF doesn't have native CSV support.
    """

    user_is_staff = request.user.groups.filter(name=REGISTRIES_VIEWER_ROLE).exists()
    if not user_is_staff:
        raise PermissionDenied()

    # Create the HttpResponse object with the appropriate CSV header.
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="registry.csv"'
    writer = csv.writer(response)
    writer.writerow(REGISTRY_EXPORT_HEADER_COLUMNS)

    queryset = ordered_person_search_qs(request)
    for row in build_row(queryset):
        writer.writerow(row)

    return response


@api_view(['GET'])
def xlsx_export_v2(request):
    """
    Export the registry as XLSX.
    """
    mime_type = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

    user_is_staff = request.user.groups.filter(name=REGISTRIES_VIEWER_ROLE).exists()
    if not user_is_staff:
        raise PermissionDenied()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(REGISTRY_EXPORT_HEADER_COLUMNS)
    for i, column_name in enumerate(REGISTRY_EXPORT_HEADER_COLUMNS):
        col_letter = get_column_letter(i + 1)
        ws.column_dimensions[col_letter].width = len(column_name)
    queryset = ordered_person_search_qs(request)
    for row in build_row(queryset):
        ws.append([str(col) if col else '' for col in row])
    response = HttpResponse(content=save_virtual_workbook(wb), content_type=mime_type)
    response['Content-Disposition'] = 'attachment; filename="registry.xlsx"'
    return response
