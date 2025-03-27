"""
    Licensed under the Apache License, Version 2.0 (the "License");
    you may not use this file except in compliance with the License.
    You may obtain a copy of the License at

        http://www.apache.org/licenses/LICENSE-2.0

    Unless required by applicable law or agreed to in writing, software
    distributed under the License is distributed on an "AS IS" BASIS,
    WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
    See the License for the specific language governing permissions and
    limitations under the License.
"""
import csv
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.writer.excel import save_virtual_workbook

from django_filters import rest_framework as restfilters
from django.http import HttpResponse, HttpResponseRedirect
from django.core.exceptions import PermissionDenied
from rest_framework import filters
from rest_framework.generics import ListAPIView
from rest_framework.decorators import api_view

from registries.permissions import RegistriesEditPermissions
from registries.models import Person

from .views import (person_search_qs, exclude_persons_without_registrations)

REGISTRY_EXPORT_HEADER_COLUMNS = [
    'person_name',
    'registration_number',
    'registration_status',
    'well_driller_orcs_number',
    'pump_installer_orcs_number',
    'description',
    'primary_certificate',
    'company_name',
    'company_address',
    'company_province_state',
    'company_tel',
    'company_fax',
    'company_email',
    'company_url',
    'contact_cell',
    'contact_tel',
    'contact_email',
    'person_guid',
    'org_guid',
]

def build_record(person, registration=None, organization=None, application=None):
    # NOTE: must be kept in sync with REGISTRY_EXPORT_HEADER_COLUMNS above
    record = {
        'person_name': person.name,
        'registration_number': registration.registration_no if registration else None,
        'registration_status': application.display_status if application else None,
        'well_driller_orcs_number': person.well_driller_orcs_no,
        'pump_installer_orcs_number': person.pump_installer_orcs_no,
        'description': application.subactivity.description if application else None,
        'primary_certificate': application.primary_certificate if application else None,
        'company_name': organization.name if organization else None,
        'company_address': organization.mailing_address if organization else None,
        'company_province_state': organization.province_state if organization else None,
        'company_tel': organization.main_tel if organization else None,
        'company_fax': organization.fax_tel if organization else None,
        'company_email': organization.email if organization else None,
        'company_url': organization.website_url if organization else None,
        'contact_cell': person.contact_cell,
        'contact_tel': person.contact_tel,
        'contact_email': person.contact_email,
        'person_guid': person.person_guid,
        'org_guid': registration.organization_id if registration else None,
    }

    return list(record.values())

def build_row(queryset):
    for person in queryset:
        registrantions = person.registrations.all()
        if len(registrantions) == 0:
            yield build_record(person)
        else:
            for registration in registrantions:
                applications = registration.applications.all()
                if len(applications) == 0:
                    yield build_record(person, registration, registration.organization)
                else:
                    for application in applications:
                        yield build_record(person, registration, registration.organization, application)


class CSVExportV2(ListAPIView):
    """
    Export the registry as CSV. This is done in a vanilla functional Django view instead
    of DRF, because DRF doesn't have native CSV support.
    """

    permission_classes = (RegistriesEditPermissions,)

    # Allow searching on name fields, names of related companies, etc.
    filter_backends = (restfilters.DjangoFilterBackend,
                       filters.SearchFilter, filters.OrderingFilter)
    ordering_fields = ('surname', 'registrations__organization__name')
    ordering = ('surname',)
    search_fields = (
        'first_name',
        'surname',
        'registrations__organization__name',
        'registrations__organization__city',
        'registrations__registration_no'
    )

    # fetch related companies and registration applications (prevent duplicate database trips)
    queryset = Person.objects.all()

    def get_queryset(self):
        """ Returns Person queryset, removing non-active and unregistered drillers for anonymous users """

        return person_search_qs(self.request)

    def get(self, request, *args, **kwargs):
        # Returns self.list - overridden for schema documentation
        return self.list(request, *args, **kwargs)

    def list(self, request, **kwargs):
        queryset = self.get_queryset()
        filtered_queryset = self.filter_queryset(queryset)
        filtered_queryset = exclude_persons_without_registrations(request, filtered_queryset) 

        # Create the HttpResponse object with the appropriate CSV header.
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="registry.csv"'
        writer = csv.writer(response)
        writer.writerow(REGISTRY_EXPORT_HEADER_COLUMNS)

        for row in build_row(filtered_queryset):
            writer.writerow(row)

        return response


class XLSXExportV2(ListAPIView):
    """
    Export the registry as XLSX.
    """
    permission_classes = (RegistriesEditPermissions,)

    # Allow searching on name fields, names of related companies, etc.
    filter_backends = (restfilters.DjangoFilterBackend,
                       filters.SearchFilter, filters.OrderingFilter)
    ordering_fields = ('surname', 'registrations__organization__name')
    ordering = ('surname',)
    search_fields = (
        'first_name',
        'surname',
        'registrations__organization__name',
        'registrations__organization__city',
        'registrations__registration_no'
    )

    # fetch related companies and registration applications (prevent duplicate database trips)
    queryset = Person.objects.all()

    def get_queryset(self):
        """ Returns Person queryset, removing non-active and unregistered drillers for anonymous users """

        return person_search_qs(self.request)

    def get(self, request, *args, **kwargs):
        # Returns self.list - overridden for schema documentation
        return self.list(request, *args, **kwargs)

    def list(self, request, **kwargs):
        queryset = self.get_queryset()
        filtered_queryset = self.filter_queryset(queryset)
        filtered_queryset = exclude_persons_without_registrations(request, filtered_queryset) 

        mime_type = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(REGISTRY_EXPORT_HEADER_COLUMNS)
        for i, column_name in enumerate(REGISTRY_EXPORT_HEADER_COLUMNS):
            col_letter = get_column_letter(i + 1)
            ws.column_dimensions[col_letter].width = len(column_name)
        for row in build_row(filtered_queryset):
            ws.append([str(col) if col else '' for col in row])
        response = HttpResponse(content=save_virtual_workbook(wb), content_type=mime_type)
        response['Content-Disposition'] = 'attachment; filename="registry.xlsx"'
        return response
