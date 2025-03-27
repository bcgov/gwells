"""
    Licensed under the Apache License, Version 2.0 (the "License");
    you may not use this file except in compliance with the License.
    You may obtain a copy of the License at

        http://www.apache.org/licenses/LICENSE-2.0

    Unless required by applicable law or agreed to in writing, software
    distributed under the License is distributed on an "AS IS" BASIS,
    WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
    See the License for the specific language governing permissions and
    limitations under the License.
"""
import tempfile

import openpyxl
from rest_framework.renderers import BaseRenderer
from rest_framework_csv.renderers import CSVStreamingRenderer


COLUMN_LABELS = {
    "well_guid": "Well Unique Id",
    "well_tag_number": "Well Tag Number",
    "identification_plate_number": "Well Identification Plate Number",
    "owner_full_name": "Owner Name",
    "well_class": "Well Class",
    "well_subclass": "Well Subclass",
    "well_status": "Well Status",
    "licenced_status": "Licenced Status",
    "street_address": "Street Address",
    "city": "Town/City",
    "legal_lot": "Lot",
    "legal_plan": "Plan",
    "legal_district_lot": "District Lot",
    "legal_block": "Block",
    "legal_section": "Section",
    "legal_township": "Township",
    "legal_range": "Range",
    "land_district": "Land District",
    "legal_pid": "Property Identification Description (PID)",
    "well_location_description": "Description of Well Location",
    "construction_start_date": "Construction Start Date",
    "construction_end_date": "Construction Date",
    "alteration_start_date": "Alteration Start Date",
    "alteration_end_date": "Alteration Date",
    "decommission_start_date": "Decommission Start Date",
    "decommission_end_date": "Decommission Date",
    "company_of_person_responsible": "Drilling Company",
    "company_of_person_responsible_name": "Drilling Company",
    "person_responsible": "Person responsible for work",
    "person_responsible_name": "Person responsible for work",
    "driller_name": "Person who completed the work",
    "well_identification_plate_attached": "Well Identification Plate Is Attached",
    "id_plate_attached_by": "Well identification plate attached by",
    "water_supply_system_name": "Water Supply System Name",
    "water_supply_system_well_name": "Water Supply System Well Name",
    "latitude": "Latitude",
    "longitude": "Longitude",
    "coordinate_acquisition_code": "Location Accuracy Code",
    "ground_elevation": "Ground Elevation",
    "ground_elevation_method": "Elevation Determined By",
    "intended_water_use": "Intended Water Use",
    "drilling_methods": "Drilling Methods",
    "well_orientation_status": "Orientation of Well",
    "surface_seal_material": "Surface Seal Material",
    "surface_seal_thickness": "Surface Seal Thickness",
    "surface_seal_method": "Surface Seal Installation Method",
    "surface_seal_depth": "Surface Seal Depth",
    "backfill_type": "Backfill Material Above Surface Seal",
    "backfill_depth": "Backfill Depth",
    "liner_material": "Liner Material",
    "liner_diameter": "Liner Diameter",
    "liner_thickness": "Liner Thickness",
    "liner_from": "Liner From",
    "liner_to": "Liner To",
    "screen_intake_method": "Intake Method",
    "screen_type": "Type",
    "screen_material": "Material",
    "other_screen_material": "Specify Other Screen Material",
    "screen_opening": "Opening",
    "screen_bottom": "Bottom",
    "other_screen_bottom": "Specify Other Screen Bottom",
    "screen_information": "Screen Information",
    "filter_pack_from": "Filter Pack From",
    "filter_pack_to": "Filter Pack To",
    "filter_pack_thickness": "Filter Pack Thickness",
    "filter_pack_material": "Filter Pack Material",
    "filter_pack_material_size": "Filter Pack Material Size",
    "development_methods": "Development Methods",
    "development_hours": "Development Total Duration",
    "development_notes": "Development Notes",
    "yield_estimation_method": "Estimation Method",
    "yield_estimation_rate": "Estimation Rate",
    "yield_estimation_duration": "Estimation Duration",
    "well_yield_unit": "Yield Unit",
    "static_level_before_test": "SWL Before Test",
    "drawdown": "Drawdown",
    "hydro_fracturing_performed": "Hydro-fracturing Performed?",
    "hydro_fracturing_yield_increase": "Well Yield Increase Due to Hydro-fracturing",
    "recommended_pump_depth": "Recommended pump depth",
    "recommended_pump_rate": "Recommended pump rate",
    "water_quality_characteristics": "Obvious Water Quality Characteristics",
    "water_quality_colour": "Water Quality Colour",
    "water_quality_odour": "Water Quality Odour",
    "total_depth_drilled": "Total Depth Drilled",
    "finished_well_depth": "Finished Well Depth",
    "well_yield": "Estimated Well Yield",
    "diameter": "Diameter",
    "observation_well_number": "Observation Well Number",
    "observation_well_status": "Observation Well Status",
    "ems": "Environmental Monitoring System (EMS) ID",
    "aquifer": "Aquifer ID Number",
    "utm_zone_code": "Zone",
    "utm_northing": "UTM Northing",
    "utm_easting": "UTM Easting",
    "bcgs_id": "BCGS Mapsheet Number",
    "decommission_reason": "Reason for Decommission",
    "decommission_method": "Method of Decommission",
    "decommission_sealant_material": "Decommission Sealant Material",
    "decommission_backfill_material": "Decommission Backfill Material",
    "decommission_details": "Decommission Details",
    "aquifer_vulnerability_index": "AVI",
    "aquifer_lithology": "Aquifer Lithology",
    "storativity": "Storativity",
    "transmissivity": "Transmissivity",
    "hydraulic_conductivity": "Hydraulic Conductivity",
    "specific_storage": "Specific Storage",
    "specific_yield": "Specific Yield",
    "testing_method": "Testing Method",
    "testing_duration": "Testing Duration",
    "analytic_solution_type": "Analytic Solution Type",
    "boundary_effect": "Boundary Effect",
    "final_casing_stick_up": "Final Casing Stick Up",
    "bedrock_depth": "Depth to Bedrock",
    "artesian_flow": "Artesian Flow",
    "artesian_pressure": "Artesian Pressure",
    "well_cap_type": "Well Cap",
    "well_disinfected_status": "Well Disinfected Code",
    "static_water_level": "Static Water Level (BTOC)",
    "artesian_pressure_head": "Artesian Pressure head",
    "artesian_conditions": "Artesian Conditions",
    "alternative_specs_submitted": "Alt specs submitted",
    "technical_report": "Technical Report",
    "drinking_water_protection_area_ind": "Drinking Water Protection Area",
    "well_activity_type": "Work/Report Type",
    "work_start_date": "Work Start Date",
    "work_end_date": "Work End Date",
    "create_user": "Created By",
    "create_date": "Created Date",
    "update_date": "Updated Date",
    "update_user": "Updated By",
    "natural_resource_region": "Natural Resource Region",
    "comments": "Comments",
    "internal_comments": "Internal Comments",
    "cross_referenced": "Cross Referenced",
    "cross_referenced_date": "Cross Referenced Date",
    "cross_referenced_by": "Cross Referenced By",
    "geocode_distance": "Geocode Distance",
    "distance_to_pid": "Distance to Matching PID",
    "score_address": "Score Address",
    "score_city": "Score City"
}


class WellListExcelRenderer(BaseRenderer):
    media_type = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    format = 'xlsx'

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)

        self.workbook = openpyxl.Workbook(write_only=True)
        self.sheet = self.workbook.create_sheet()
        self.headers = []

    def write_headers(self, headers):
        self.headers = headers

        header_cells = []
        for index, header_key in enumerate(self.headers, 1):
            label = COLUMN_LABELS.get(header_key)

            column = openpyxl.utils.get_column_letter(index)
            width = (len(label) + 2) * 1.2
            self.sheet.column_dimensions[column].width = width

            cell = openpyxl.cell.WriteOnlyCell(self.sheet, value=label)
            cell.font = openpyxl.styles.Font(bold=True)
            header_cells.append(cell)

        self.sheet.append(header_cells)

    def write_row(self, data):
        row = []
        for key in self.headers:
            cell_value = data.get(key, None)
            cell = openpyxl.cell.WriteOnlyCell(self.sheet, value=cell_value)
            row.append(cell)

        self.sheet.append(row)

    def render(self, data, media_type=None, renderer_context=None):
        """
        Renders serialized data into Excel using openpyxl.

        We return an open temporary file handle a here, so we can use Django's
        FileResponse, since pyopenxml doesn't actually support streams.
        """
        if renderer_context is None:
            renderer_context = {}

        # Don't allow any columns we haven't defined in COLUMN_LABELS
        initial_headers = renderer_context.get('header', [])
        if initial_headers:
            headers = [
                header for header in initial_headers
                if header in COLUMN_LABELS.keys()
            ]
        else:
            headers = COLUMN_LABELS.keys()

        self.write_headers(headers)

        for row in data:
            self.write_row(row)

        outfile = tempfile.NamedTemporaryFile(delete=True)
        self.workbook.save(outfile.name)

        return outfile


class WellListCSVRenderer(CSVStreamingRenderer):
    header = COLUMN_LABELS.keys()
    labels = COLUMN_LABELS

    def render(self, data, media_type=None, renderer_context=None):
        """Sanitize any header values we get from the renderer context
        """
        if renderer_context is None:
            renderer_context = {}

        # Don't allow any columns we haven't defined in COLUMN_LABELS
        initial_headers = renderer_context.get('header', [])
        if initial_headers:
            renderer_context['header'] = [
                header for header in initial_headers
                if header in COLUMN_LABELS.keys()
            ]

        return super().render(data,
                              media_type=media_type,
                              renderer_context=renderer_context)
