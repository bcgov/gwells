"""
    Licensed under the Apache License, Version 2.0 (the "License");
    you may not use this file except in compliance with the License.
    You may obtain a copy of the License at

        http://www.apache.org/licenses/LICENSE-2.0

    Unless required by applicable law or agreed to in writing, software
    distributed under the License is distributed on an "AS IS" BASIS,
    WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
    See the License for the specific language governing permissions and
    limitations under the License.
"""
import csv
import zipfile
import os
import logging
import string

from django.core.management.base import BaseCommand
from django.db import connection

from minio import Minio
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
from openpyxl.worksheet.write_only import WriteOnlyCell

from gwells.settings.base import get_env_variable
from gwells.management.commands import ResultIter

# Run from command line :
# python manage.py export
#
# For development/debugging, it's useful to skip upload and cleanup
# python manage.py export --cleanup=0 --upload=0

logger = logging.getLogger(__name__)

"""
Well v1 & v2
"""
WELL_EXPORT_VIEW_V1 = ("""
    select * from export_well_v1_view
""")
WELL_EXPORT_VIEW_V2 = ("""
    select * from export_well_v2_view
""")

"""
Lithology v1 & v2
"""
LITHOLOGY_EXPORT_VIEW_V1 = ("""
    select * from export_well_lithology_v1_view
""")
LITHOLOGY_EXPORT_VIEW_V2 = ("""
    select * from export_well_lithology_v2_view
""")

"""
Casing v1 & v2
"""
CASING_EXPORT_VIEW_V1 = ("""
    select * from export_well_casing_v1_view
""")
CASING_EXPORT_VIEW_V2 = ("""
    select * from export_well_casing_v2_view
""")

"""
Screen v1 & v2
"""
SCREEN_EXPORT_VIEW_V1 = ("""
    select * from export_well_screen_v1_view
""")
SCREEN_EXPORT_VIEW_V2 = ("""
    select * from export_well_screen_v2_view
""")

"""
Perforation v1 & v2
"""
PERFORATION_EXPORT_VIEW_V1 = ("""
    select * from export_well_perforation_v1_view 
""")
PERFORATION_EXPORT_VIEW_V2 = ("""
    select * from export_well_perforation_v2_view 
""")

"""
Drilling Method v1 & v2
    Note: v1 and v2 are the same
"""
DRILLING_METHOD_EXPORT_VIEW_V1 = ("""
    select * from export_well_drilling_v1_view
""")
DRILLING_METHOD_EXPORT_VIEW_V2 = DRILLING_METHOD_EXPORT_VIEW_V1

"""
Development Method v1 & v2
    Note: v1 and v2 are the same
"""
DEVELOPMENT_METHOD_EXPORT_VIEW_V1 = ("""
    select * from export_well_development_v1_view
""")
DEVELOPMENT_METHOD_EXPORT_VIEW_V2 = DEVELOPMENT_METHOD_EXPORT_VIEW_V1

"""
Stats v1 & v2
"""
STATS_SQL_V1 = ("""
select
    now()::timestamp(0) as date_time_created_utc,
    'The Groundwater Wells and Aquifers application may not include all groundwater wells as registration with the ' ||
    'Province was voluntary until February 2016. ' ||
    'The data provided has been collected/uploaded by the public and is provided on an "as is" basis, and as such ' ||
    'may impact search results and be subject to change.' as disclaimer_txt
""")
STATS_SQL_V2 = STATS_SQL_V1

"""
Aquifer Parameters v1 & v2
"""
AQUIFER_PARAMETERS_SQL_V1 = ("""
    select * from export_aquifer_parameters_v1_view
""")
AQUIFER_PARAMETERS_SQL_V2 = AQUIFER_PARAMETERS_SQL_V1


class Command(BaseCommand):

    def __init__(self):
        """
        Declare our class variables
        Manage the versioning descriptor
        Version1 is blank, this is where the original file was stored.
        We're keeping the default location so unknown consumers of this file can
            continue to access it
        """
        super().__init__()
        self.version1 = ''
        self.version2 = 'v2'

        # Create a dictionary to map sql statements to spreadsheet worksheets.
        self.versioning_descriptor = [
            {
                'version': self.version1,
                'sheets_sql': {
                    'well': WELL_EXPORT_VIEW_V1,
                    'lithology': LITHOLOGY_EXPORT_VIEW_V1,
                    'casing': CASING_EXPORT_VIEW_V1,
                    'screen': SCREEN_EXPORT_VIEW_V1,
                    'perforation': PERFORATION_EXPORT_VIEW_V1,
                    'drilling_method': DRILLING_METHOD_EXPORT_VIEW_V1,
                    'development_method': DEVELOPMENT_METHOD_EXPORT_VIEW_V1,
                    'pt_aquifer_parameters': AQUIFER_PARAMETERS_SQL_V1,
                    'stats': STATS_SQL_V1
                }
            },
            {
                'version': self.version2,
                'sheets_sql': {
                    'well': WELL_EXPORT_VIEW_V2,
                    'lithology': LITHOLOGY_EXPORT_VIEW_V2,
                    'casing': CASING_EXPORT_VIEW_V2,
                    'screen': SCREEN_EXPORT_VIEW_V2,
                    'perforation': PERFORATION_EXPORT_VIEW_V2,
                    'drilling_method': DRILLING_METHOD_EXPORT_VIEW_V2,
                    'development_method': DEVELOPMENT_METHOD_EXPORT_VIEW_V2,
                    'pt_aquifer_parameters': AQUIFER_PARAMETERS_SQL_V2,
                    'stats': STATS_SQL_V2
                }
            }
        ]

        self.worksheet_styles = [{'worksheet_name': 'stats', 'column_name': 'disclaimer_txt', 'column_width': 240}]

    def add_arguments(self, parser):
        """
        Arguments added for debugging purposes.

        :param parser: the django handled arguments parser
        :Example: don't cleanup, don't upload: python manage.py export --cleanup=0 --upload=0
        :Example: cleanup, don't upload: python manage.py export --cleanup=1 --upload=0
        :Example: cleanup and upload: python manage.py export --cleanup=1 --upload=1
        """
        parser.add_argument('--cleanup', type=int, nargs='?', help='If 1, remove file when done', default=1)
        parser.add_argument('--upload', type=int, nargs='?', help='If 1, upload the file', default=1)

    def handle(self, *args, **options):
        """
        Entry point for Django Command.
        for each version (v1='', v2='v2'), generate our outputs,
           upload our files if option is set,
           cleanup our filesystem if option is set

        :param args: django managed args
        :param options: django managed options
        """
        logger.info('starting export')
        zip_filename = 'gwells.zip'
        spreadsheet_filename = 'gwells.xlsx'
        for version_desc in self.versioning_descriptor:
            version = version_desc['version']
            sheets = version_desc['sheets_sql']
            self.generate_files(zip_filename, spreadsheet_filename, sheets)
            if options['upload'] == 1:
                self.upload_files(zip_filename, spreadsheet_filename, version)
            if options['cleanup'] == 1:
                logger.info('cleaning up')
                for filename in (zip_filename, spreadsheet_filename):
                    if os.path.exists(filename):
                        os.remove(filename)

        logger.info('export complete')
        self.stdout.write(self.style.SUCCESS('export complete'))

    def upload_files(self, zip_filename, spreadsheet_filename, version: str):
        """
        Upload files to S3 bucket.

        :param zip_filename: the filename of the zip file
        :param spreadsheet_filename: the filename of the spreadsheet
        :param version: the version to use
        """
        is_secure = get_env_variable('S3_USE_SECURE', '1', warn=False) is '1'
        minio_client = Minio(get_env_variable('S3_HOST'),
                             access_key=get_env_variable('S3_PUBLIC_ACCESS_KEY'),
                             secret_key=get_env_variable('S3_PUBLIC_SECRET_KEY'),
                             secure=is_secure)
        for filename in (zip_filename, spreadsheet_filename):
            logger.info('uploading {}'.format(filename))
            with open(filename, 'rb') as file_data:
                file_stat = os.stat(filename)
                # our target supports versioned location writing, if the version is blank, continue
                #   outputting just as we have in the past, otherwise output to export/versionNumberHere/fileNameHere
                target = f'export/{filename}' if version == '' else f'export/{version}/{filename}'
                minio_client.put_object(get_env_variable('S3_WELL_EXPORT_BUCKET'),
                                        target,
                                        file_data,
                                        file_stat.st_size)

    def export(self, workbook, gwells_zip, worksheet_name, cursor):
        """
        Generates the csv/zip file content by taking the using
            the cursor and iterating over the resultset.

        :param workbook: openpyxl.workbook.workbook instance
        :param gwells_zip: file handle to zip
        :param worksheet_name: the name of the worksheet in the workbook
        :param cursor: a cursor from django database wrapper
        """
        logger.info('exporting {}'.format(worksheet_name))
        worksheet = workbook.create_sheet(worksheet_name)
        csv_file = '{}.csv'.format(worksheet_name)
        # If any of the export files already exist, delete them
        if os.path.exists(csv_file):
            os.remove(csv_file)
        with open(csv_file, 'w') as csvfile:
            csvwriter = csv.writer(csvfile, dialect='excel')

            values = []
            cells = []
            # Write the headings
            for index, field in enumerate(cursor.description):
                field_name = field[0]
                values.append(field_name)
                cell = WriteOnlyCell(worksheet, value=field_name)
                cell.font = Font(bold=True)
                cells.append(cell)
            columns = len(values)

            for index, value in enumerate(values):
                # style modifications can be applied to outputs (haven't been able to get wrap_text working), feels
                #   like it's being overwritten by the value output section below (Write the values).
                # For now this is a decent solution, we can apply style modifications to a given column
                #   based on our provided configuration
                style_applied = False
                for worksheet_style in self.worksheet_styles:
                    if worksheet_name == worksheet_style['worksheet_name'] and value == worksheet_style['column_name']:
                        worksheet.column_dimensions[get_column_letter(index + 1)].width = worksheet_style['column_width']
                        style_applied = True
                        break
                if not style_applied:
                    worksheet.column_dimensions[get_column_letter(index + 1)].width = len(value) + 2

            worksheet.append(cells)
            csvwriter.writerow(values)

            # Write the values
            row_index = 0
            for row, record in enumerate(ResultIter(cursor)):
                values = []
                num_values = 0
                for col, value in enumerate(record):
                    if not (value == "" or value is None):
                        num_values += 1
                    if type(value) is str:
                        # There are lots of non-printable characters in the source data that can cause
                        # issues in the export, so we have to clear them out.
                        v = ''.join([s for s in value if s in string.printable])
                        # We can't have something starting with an = sign,
                        # it would be interpreted as a formula in excel.
                        if v.startswith('='):
                            v = '\'{}'.format(v)
                        values.append(v)
                    else:
                        values.append(value)
                if num_values > 1:
                    # We always have a well_tag_number, but if that's all we have, then just skip this record
                    row_index += 1
                    csvwriter.writerow(values)
                    worksheet.append(values)

            filter_reference = 'A1:{}{}'.format(get_column_letter(columns), row_index + 1)
            worksheet.auto_filter.ref = filter_reference

        gwells_zip.write(csv_file)
        if os.path.exists(csv_file):
            # After adding the csv file to the zip, delete it.
            os.remove(csv_file)

    def generate_files(self, zip_filename, spreadsheet_filename, sheets: dict):
        """
        Runs sql statements passing the cursor to the self.export method.
        Creates the spreadsheet and zip file outputs

        :param zip_filename: a string filename to use
        :param spreadsheet_filename: a string filename to use
        :param sheets: the sheet/sql map
        """
        # If there is an existing zip file, remove it.
        if os.path.exists(zip_filename):
            os.remove(zip_filename)
        with zipfile.ZipFile(zip_filename, 'w', compression=zipfile.ZIP_DEFLATED) as gwells_zip:
            if os.path.exists(spreadsheet_filename):
                os.remove(spreadsheet_filename)
            workbook = Workbook(write_only=True)

            for sheet, sql in sheets.items():
                logger.info('creating {} cursor'.format(sheet))
                with connection.cursor() as cursor:
                    cursor.execute(sql)
                    self.export(workbook, gwells_zip, sheet, cursor)
            workbook.save(filename=spreadsheet_filename)
            # Add a readme to the zipfile.
            arcname = 'README.md'
            readme = os.path.join(os.path.dirname(__file__), arcname)
            gwells_zip.write(readme, arcname)
